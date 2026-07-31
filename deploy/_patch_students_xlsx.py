#!/usr/bin/env python3
"""Download Students.xlsx from VPS, add visitor columns, upload back."""
from __future__ import annotations

import os
from pathlib import Path

import paramiko

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    raise SystemExit("pip install openpyxl")

ROOT = Path(r"C:\xampp7\htdocs\Xander-school")
LOCAL = ROOT / "public" / "assets" / "templates" / "Students.xlsx"
HOST = "66.29.135.120"
USER = "root"
PASSWORD = os.environ.get("VPS_PASSWORD", "6W7sa2g4dMEwcN80ZU")
REMOTE = "/opt/xander-school/app/public/assets/templates/Students.xlsx"

VISITOR_HEADERS = [
    "Visitor1 Name",
    "Visitor1 Phone",
    "Visitor1 Relationship",
    "Visitor2 Name",
    "Visitor2 Phone",
    "Visitor2 Relationship",
]


def patch_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start_col = 15  # O
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for i, title in enumerate(VISITOR_HEADERS):
        col = start_col + i
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    # Example row 2 hints
    ws.cell(row=2, column=15, value="JEAN BAPTISTE NAM")
    ws.cell(row=2, column=16, value="0788000000")
    ws.cell(row=2, column=17, value="Father")
    ws.cell(row=2, column=18, value="MARIE CLAIRE NAM")
    ws.cell(row=2, column=19, value="0788111111")
    ws.cell(row=2, column=20, value="Mother")
    wb.save(path)
    print("Patched", path)


def main() -> None:
    LOCAL.parent.mkdir(parents=True, exist_ok=True)
    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(HOST, username=USER, password=PASSWORD, timeout=90)
    sftp = c.open_sftp()
    try:
        sftp.get(REMOTE, str(LOCAL))
        print("Downloaded", REMOTE)
    except FileNotFoundError:
        print("Remote missing — creating fresh template")
        wb = openpyxl.Workbook()
        ws = wb.active
        base = [
            "First Name", "Last Name", "Reg No (Leave it to auto generate)", "Gender",
            "Birth Date (Ex: 2000-01-25)", "Mode", "Nationality", "Father Names",
            "Phone Number", "Mother Names", "Phone Number2", "Guardian Names",
            "Phone Number3", "Religion",
        ]
        for i, h in enumerate(base, 1):
            ws.cell(row=1, column=i, value=h)
        wb.save(LOCAL)
    sftp.close()
    c.close()

    patch_workbook(LOCAL)

    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(HOST, username=USER, password=PASSWORD, timeout=90)
    sftp = c.open_sftp()
    sftp.put(str(LOCAL), REMOTE)
    sftp.close()
    c.close()
    print("Uploaded to VPS")


if __name__ == "__main__":
    main()

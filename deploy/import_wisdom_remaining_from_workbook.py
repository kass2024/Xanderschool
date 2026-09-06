#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import os
import re
import shlex
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import paramiko

ROOT = Path(r"C:\xampp7\htdocs\Xander-school")
DEFAULT_WORKBOOK = Path(r"C:\Users\user\Downloads\TEACHERS_ALLOCATION_COMPILED.xlsx")
HOST = "66.29.135.120"
USER = "root"
PASSWORD = os.environ.get("VPS_PASSWORD", "6W7sa2g4dMEwcN80ZU")
MYSQL_CMD = "docker exec -i xander_school_mysql mysql -uxander_school -pXsApp_3hT6yU1bC5nM9 --batch --raw --skip-column-names iotxa_db -e"

SCHOOL_ID = 27
ACADEMIC_YEAR_ID = 16
ALL_TERMS = "1,2,3"
TEACHER_POST = 2
SHIFT_ID = 24
CREATED_BY = 1
DEFAULT_PASSWORD = "Teacher@2026"
CREATE_SOURCE = "xlsx_import"

TVET_SUBJECTS = {
    "COMPUTER SCIENCE",
    "CRATE A BUSINESS",
    "DEV GAME IN VUE",
    "FINANCCIAL ACC",
    "MANAGT ACCOUNTING",
    "MCH",
    "FN",
    "F.N",
    "PHARMACOLOGY",
    "MATH&ROBOTICS CHALLENGE",
}

SPECIAL_SUBJECTS = {
    "F.N",
    "FN",
    "CLINICAL ATTACHMENT",
    "MEDICAL PATHOLOGY",
    "MAEDICAL PATHOLOGY",
    "SURGICAL PATHOLOGY",
}

NON_EXAM_SUBJECTS = {
    "CHINESE",
    "FRENCH",
    "KISWAHILI",
    "SPORT",
}

SUBJECT_TITLES = {
    "BIOLOGY": "Biology",
    "CHEM": "Chemistry",
    "CHINESE": "Chinese",
    "CLINICAL ATTACHMENT": "Clinical Attachment",
    "COMPUTER SCIENCE": "Computer Science",
    "CRATE A BUSINESS": "Create a Business",
    "DEV GAME IN VUE": "Dev Game in Vue",
    "ENGLISH": "English",
    "ENTREPRENEURSHIP": "Entrepreneurship",
    "F.N": "F.N",
    "FN": "F.N",
    "FINANCCIAL ACC": "Financial Accounting",
    "FRENCH": "French",
    "GEOGRAPHY": "Geography",
    "HISTORY": "History",
    "ICT": "ICT",
    "KINYARWANDA": "Kinyarwanda",
    "KISWAHILI": "Kiswahili",
    "MANAGT ACCOUNTING": "Management Accounting",
    "MATHEMATICS": "Mathematics",
    "MAEDICAL PATHOLOGY": "Medical Pathology",
    "MATH&ROBOTICS CHALLENGE": "Math & Robotics Challenge",
    "MEDICAL PATHOLOGY": "Medical Pathology",
    "MCH": "MCH",
    "PHARMACOLOGY": "Pharmacology",
    "PHYSICS": "Physics",
    "SPORT": "Sport",
    "SURGICAL PATHOLOGY": "Surgical Pathology",
}

SUBJECT_BASE_CODES = {
    "BIOLOGY": "BIO",
    "CHEM": "CHEM",
    "CHINESE": "CHIN",
    "COMPUTER SCIENCE": "CS",
    "CRATE A BUSINESS": "CAB",
    "DEV GAME IN VUE": "DGV",
    "ENGLISH": "ENG",
    "ENTREPRENEURSHIP": "ENT",
    "F.N": "FN",
    "FN": "FN",
    "FINANCCIAL ACC": "FINA",
    "FRENCH": "FRE",
    "GEOGRAPHY": "GEO",
    "HISTORY": "HIS",
    "ICT": "ICT",
    "KINYARWANDA": "KINY",
    "KISWAHILI": "KISW",
    "MANAGT ACCOUNTING": "MGA",
    "MATHEMATICS": "MATH",
    "MATH&ROBOTICS CHALLENGE": "MATHROBO",
    "MCH": "MCH",
    "PHARMACOLOGY": "PHARM",
    "PHYSICS": "PHY",
    "SPORT": "SPORT",
}

NAME_ALIASES = {
    "BUA SARAFNAH": ["BUA SARAFINAH"],
    "BUA SARAFINAH": ["BUA SARAFNAH"],
    "AYOBIKEKA BENOIT": ["AYEBIKEKA BENOIT"],
    "AYEBIKEKA BENOIT": ["AYOBIKEKA BENOIT"],
    "RWANGA BENJANIN": ["RWANGA BENJAMIN"],
    "RWANGA BENJAMIN": ["RWANGA BENJANIN"],
    "CIBALONZA": ["MUSANGANYA CIBALONZA"],
    "ANDREW": ["NITIRUSHWAMABOKO ANDREW"],
}


@dataclass
class TeacherEntry:
    teacher: str
    subject: str
    phone: str
    email: str
    classes: list[tuple[str, int]]


def mysql_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def sql_value(value: object) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + mysql_escape(str(value)) + "'"


def normalize_name(name: str) -> str:
    name = re.sub(r"\(.*?\)", "", str(name or ""))
    name = re.sub(r"[^A-Za-z0-9]+", " ", name).strip().upper()
    return re.sub(r"\s+", " ", name)


def name_tokens(name: str) -> set[str]:
    return {p for p in normalize_name(name).split(" ") if p}


def names_match(a: str, b: str) -> bool:
    na = normalize_name(a)
    nb = normalize_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if nb in NAME_ALIASES.get(na, []) or na in NAME_ALIASES.get(nb, []):
        return True
    ta = name_tokens(na)
    tb = name_tokens(nb)
    return bool(ta and tb and (ta <= tb or tb <= ta))


def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D+", "", str(phone or ""))
    if not digits:
        return ""
    if digits.startswith("250") and len(digits) == 12:
        return "0" + digits[3:]
    if digits.startswith("7") and len(digits) == 9:
        return "0" + digits
    return digits


def normalize_email(email: str) -> str:
    return str(email or "").strip().lower()


def parse_person_name(display_name: str) -> tuple[str, str]:
    cleaned = normalize_name(display_name)
    parts = cleaned.split()
    if not parts:
        return ("UNKNOWN", "TEACHER")
    if len(parts) == 1:
        return (parts[0], "TEACHER")
    return (" ".join(parts[:-1]), parts[-1])


def email_slug(fname: str, lname: str) -> str:
    return (re.sub(r"[^a-z0-9]+", ".", f"{fname}.{lname}".lower()).strip(".") or "teacher")


def subject_title(subject: str) -> str:
    return SUBJECT_TITLES.get(normalize_name(subject), normalize_name(subject).title())


def subject_program(subject: str) -> str:
    return "tvet" if normalize_name(subject) in {normalize_name(s) for s in TVET_SUBJECTS} else "reb"


def subject_code(subject: str, credit: int, program: str) -> str:
    base = SUBJECT_BASE_CODES.get(normalize_name(subject), re.sub(r"[^A-Z0-9]+", "", normalize_name(subject))[:8] or "SUBJ")
    prefix = "SP" if program == "special" else ("TV" if program == "tvet" else "RB")
    return f"{prefix}-{base}{credit}"


def category_title(subject: str) -> str:
    return "Non-Examinable Subjects" if normalize_name(subject) in NON_EXAM_SUBJECTS else "Examinable Subjects"


def workbook_entries(path: Path) -> list[TeacherEntry]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    review = wb["Contact Match Review"]

    review_map: dict[str, dict[str, str]] = {}
    for row in review.iter_rows(values_only=True):
        status, teacher, phone, email, _note = [(str(v).strip() if v is not None else "") for v in row[:5]]
        if not teacher or status.upper() == "STATUS":
            continue
        review_map[normalize_name(teacher)] = {"phone": normalize_phone(phone), "email": normalize_email(email)}

    entries_map: dict[tuple[str, str], TeacherEntry] = {}
    current_teacher = ""
    current_phone = ""
    current_email = ""
    current_subject = ""
    started = False
    for row_idx in range(1, ws.max_row + 1):
        vals = [(str(ws.cell(row_idx, col).value).strip() if ws.cell(row_idx, col).value is not None else "") for col in range(1, 9)]
        if vals[0] == "S/N":
            started = True
            continue
        if not started:
            continue
        if vals[1]:
            review_info = review_map.get(normalize_name(vals[1]), {})
            current_teacher = vals[1]
            current_phone = review_info.get("phone") or normalize_phone(vals[6])
            current_email = review_info.get("email") or normalize_email(vals[7])
            current_subject = ""
        if vals[2]:
            current_subject = subject_title(vals[2])
        elif not current_subject:
            continue
        if not vals[2] and not vals[1] and vals[5] and row_idx < ws.max_row and ws.cell(row_idx + 1, 2).value is not None:
            current_subject = ""
            continue
        if not current_teacher or not current_subject or not vals[3]:
            continue
        key = (normalize_name(current_teacher), normalize_name(current_subject))
        entry = entries_map.get(key)
        if entry is None:
            entry = TeacherEntry(
                teacher=current_teacher,
                subject=current_subject,
                phone=current_phone,
                email=current_email,
                classes=[],
            )
            entries_map[key] = entry
        periods = int(float(vals[4])) if vals[4] else 0
        entry.classes.append((vals[3], periods))
    return list(entries_map.values())


class RemoteDb:
    def __init__(self, host: str, user: str, password: str) -> None:
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.client.connect(host, username=user, password=password, timeout=90, banner_timeout=90)

    def close(self) -> None:
        self.client.close()

    def query_rows(self, sql: str, columns: list[str]) -> list[dict[str, str]]:
        command = f"{MYSQL_CMD} {shlex.quote(sql)}"
        _stdin, stdout, stderr = self.client.exec_command(command, timeout=120)
        out = stdout.read().decode("utf-8", "replace")
        err = stderr.read().decode("utf-8", "replace")
        code = stdout.channel.recv_exit_status()
        if code != 0:
            raise RuntimeError(err or out or f"mysql query failed: {code}")
        reader = csv.reader(io.StringIO(out), delimiter="\t")
        return [{columns[i]: (row[i] if i < len(row) else "") for i in range(len(columns))} for row in reader if row]

    def execute(self, sql: str) -> None:
        command = f"{MYSQL_CMD} {shlex.quote(sql)}"
        _stdin, stdout, stderr = self.client.exec_command(command, timeout=120)
        out = stdout.read().decode("utf-8", "replace")
        err = stderr.read().decode("utf-8", "replace")
        code = stdout.channel.recv_exit_status()
        if code != 0:
            raise RuntimeError(err or out or f"mysql execute failed: {code}")

    def execute_many(self, statements: list[str], chunk_size: int = 120) -> None:
        for i in range(0, len(statements), chunk_size):
            chunk = statements[i:i + chunk_size]
            if chunk:
                self.execute(";\n".join(chunk) + ";")


def php_password_hash(password: str) -> str:
    result = subprocess.run(["php", "-r", f'echo password_hash("{password}", PASSWORD_DEFAULT);'], capture_output=True, text=True, cwd=str(ROOT))
    out = (result.stdout or "").strip()
    if result.returncode != 0 or not out:
        raise RuntimeError("failed to generate password hash")
    return out


def load_classes(db: RemoteDb) -> list[dict[str, str]]:
    return db.query_rows(
        """
        SELECT c.id, l.title, TRIM(IFNULL(c.title,'')), d.code, d.title, f.type,
               (SELECT COUNT(*) FROM class_records cr WHERE cr.class=c.id AND cr.status=1)
        FROM classes c
        JOIN levels l ON l.id = c.level
        JOIN departments d ON d.id = c.department
        JOIN faculty f ON f.id = d.faculty_id
        WHERE c.school_id = 27
        ORDER BY l.title, d.code, c.title
        """,
        ["id", "level_name", "stream", "dept_code", "dept_title", "faculty_type", "students"],
    )


def load_staff(db: RemoteDb) -> list[dict[str, str]]:
    return db.query_rows(
        "SELECT id, fname, lname, IFNULL(phone,''), IFNULL(email,''), IFNULL(post,''), IFNULL(shift_id,'') FROM staffs WHERE school_id=27 ORDER BY id",
        ["id", "fname", "lname", "phone", "email", "post", "shift_id"],
    )


def load_categories(db: RemoteDb) -> list[dict[str, str]]:
    return db.query_rows("SELECT id, title FROM course_category WHERE school_id=27 ORDER BY id", ["id", "title"])


def load_courses(db: RemoteDb) -> list[dict[str, str]]:
    return db.query_rows(
        "SELECT id, code, title, IFNULL(credit,''), IFNULL(marks,''), IFNULL(program_type,''), IFNULL(category,0), IFNULL(create_source,'') FROM courses WHERE school_id=27 ORDER BY id",
        ["id", "code", "title", "credit", "marks", "program_type", "category", "create_source"],
    )


def load_assignments(db: RemoteDb) -> list[dict[str, str]]:
    return db.query_rows(
        f"SELECT id, course, lecturer, class, IFNULL(term,'') FROM course_records WHERE year={ACADEMIC_YEAR_ID}",
        ["id", "course", "lecturer", "class", "term"],
    )


def ensure_meta_columns(db: RemoteDb) -> None:
    fields = {row["field"] for row in db.query_rows("SHOW COLUMNS FROM courses", ["field", "type", "null", "key", "default", "extra"])}
    stmts = []
    if "program_type" not in fields:
        stmts.append("ALTER TABLE courses ADD COLUMN program_type varchar(16) NOT NULL DEFAULT 'tvet' AFTER marks")
    if "create_source" not in fields:
        stmts.append("ALTER TABLE courses ADD COLUMN create_source varchar(16) NOT NULL DEFAULT 'manual' AFTER program_type")
    if stmts:
        db.execute_many(stmts, chunk_size=2)


def ensure_category(db: RemoteDb, cache: dict[str, int], title: str, dry_run: bool) -> int:
    if title in cache:
        return cache[title]
    if dry_run:
        cache[title] = -1
        return -1
    db.execute(f"INSERT INTO course_category (school_id, title, status) VALUES ({SCHOOL_ID}, {sql_value(title)}, 1)")
    fresh = load_categories(db)
    cache.update({row["title"]: int(row["id"]) for row in fresh})
    return cache[title]


def canonical_class_map(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    out: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        level = row["level_name"].strip().upper()
        stream = row["stream"].strip().upper()
        dept = row["dept_code"].strip().upper()
        key = ""
        if level in {"S1", "S2", "S3"} and stream in {"A", "B", "C"}:
            key = f"{level}{stream}"
        elif level in {"S4", "S5"} and dept in {"ST1", "ST2", "ACC", "ANP"}:
            key = f"{level}{dept}"
        elif level == "S6" and dept in {"GE", "ANP", "MPC", "MCB", "MEG", "MPG", "PCB", "PCM", "STR", "MCE"}:
            key = f"{level}{dept}"
        elif level in {"LEVEL 3", "LEVEL 4", "LEVEL 5"} and dept == "SOD":
            key = level.replace(" ", "") + dept
        if key:
            row = dict(row)
            row["_students"] = int(row["students"] or 0)
            out[key].append(row)
    return out


def pick(keys: list[str], classes: dict[str, list[dict[str, str]]], include_empty: bool) -> list[int]:
    out: list[int] = []
    seen: set[int] = set()
    for key in keys:
        rows = sorted(classes.get(key, []), key=lambda r: (-int(r["_students"]), int(r["id"])))
        if not rows:
            continue
        rows = rows if include_empty else [r for r in rows if int(r["_students"]) > 0] or [rows[0]]
        for row in rows:
            cid = int(row["id"])
            if cid not in seen:
                seen.add(cid)
                out.append(cid)
    return out


def classes_for_label(label: str, subject: str, classes: dict[str, list[dict[str, str]]]) -> list[int]:
    raw = normalize_name(label).replace(" ", "")
    special = normalize_name(subject) in {normalize_name(s) for s in SPECIAL_SUBJECTS}
    direct = {"S1A", "S1B", "S1C", "S2A", "S2B", "S2C", "S3A", "S3B", "S3C", "S4ACC", "S5ACC", "S4ANP", "S5ANP", "S6ANP", "S6GE", "LEVEL3SOD", "LEVEL4SOD", "LEVEL5SOD"}
    if raw in direct:
        return pick([raw], classes, include_empty=True)
    alias_map = {
        "S4SOD": ["LEVEL4SOD"],
        "S5SOD": ["LEVEL5SOD"],
        "S4STREAMI": ["S4ST1"],
        "S4STREAM1": ["S4ST1"],
        "S4STREAMII": ["S4ST2"],
        "S4STREMII": ["S4ST2"],
        "S4STREAM2": ["S4ST2"],
        "S5STREAMI": ["S5ST1"],
        "S5STREAM1": ["S5ST1"],
        "S5STREAMII": ["S5ST2"],
        "S5STREAM2": ["S5ST2"],
        "S4STREAMIII": ["S4ST1", "S4ST2"],
        "S4STREAMIANDII": ["S4ST1", "S4ST2"],
        "S4SREAMIANDII": ["S4ST1", "S4ST2"],
        "S5STREAMIII": ["S5ST1", "S5ST2"],
        "S5STREAMIANDII": ["S5ST1", "S5ST2"],
        "S5SREAMIANDII": ["S5ST1", "S5ST2"],
        "S6ANPANDGE": ["S6ANP", "S6GE"],
        "S6ANPANDPCB": ["S6ANP", "S6PCB"],
        "LEVEL345SOD": ["LEVEL3SOD", "LEVEL4SOD", "LEVEL5SOD"],
        "SOD": ["LEVEL3SOD", "LEVEL4SOD", "LEVEL5SOD"],
    }
    if raw in alias_map:
        return pick(alias_map[raw], classes, include_empty=True)
    if raw == "S1S6":
        return pick(["S1A", "S1B", "S1C", "S2A", "S2B", "S2C", "S3A", "S3B", "S3C", "S4ACC", "S4ST1", "S4ST2", "S5ACC", "S5ST1", "S5ST2", "S6GE", "S6STR", "S6MCE", "S6PCB", "S6MCB", "S6MPC", "S6MEG", "S6MPG", "S6PCM"], classes, include_empty=False)
    if raw == "S2":
        return pick(["S2A", "S2B", "S2C"], classes, include_empty=False)
    if raw == "S3":
        return pick(["S3A", "S3B", "S3C"], classes, include_empty=False)
    if raw == "S4":
        return pick(["S4ANP"] if special else ["S4ACC", "S4ST1", "S4ST2"], classes, include_empty=False)
    if raw == "S5":
        return pick(["S5ANP"] if special else ["S5ACC", "S5ST1", "S5ST2"], classes, include_empty=False)
    if raw == "S6":
        return pick(["S6ANP"] if special else ["S6GE", "S6STR", "S6MCE", "S6PCB", "S6MCB", "S6MPC", "S6MEG", "S6MPG", "S6PCM"], classes, include_empty=False)
    return []


def label_specificity(label: str) -> int:
    raw = normalize_name(label).replace(" ", "")
    if raw in {"S1A", "S1B", "S1C", "S2A", "S2B", "S2C", "S3A", "S3B", "S3C", "S4ACC", "S5ACC", "S4ANP", "S5ANP", "S6ANP", "S6GE", "LEVEL3SOD", "LEVEL4SOD", "LEVEL5SOD"}:
        return 100
    if raw in {"S4STREAMI", "S4STREAMII", "S5STREAMI", "S5STREAMII", "S4STREMII"}:
        return 90
    if "AND" in raw or raw.endswith("III") or raw == "S1S6" or raw == "LEVEL345SOD" or raw == "SOD":
        return 40
    if raw in {"S2", "S3", "S4", "S5", "S6"}:
        return 30
    return 70


def choose_better_candidate(current: dict[str, object], challenger: dict[str, object]) -> dict[str, object]:
    cur_spec = int(current.get("specificity", 0))
    new_spec = int(challenger.get("specificity", 0))
    if new_spec != cur_spec:
        return challenger if new_spec > cur_spec else current
    cur_load = int(current.get("subject_load", 0))
    new_load = int(challenger.get("subject_load", 0))
    if new_load != cur_load:
        return challenger if new_load > cur_load else current
    return current


def build_course_maps(rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[tuple[str, str, int], dict[str, str]]]:
    by_code: dict[str, dict[str, str]] = {}
    by_key: dict[tuple[str, str, int], dict[str, str]] = {}
    for row in rows:
        if row["code"]:
            by_code[row["code"]] = row
        if row["code"].startswith("RB-") or row["code"].startswith("TV-") or row["code"].startswith("SP-") or row.get("create_source") == CREATE_SOURCE:
            by_key[(normalize_name(row["title"]), row["program_type"] or "tvet", int(float(row["credit"] or 0)))] = row
    return by_code, by_key


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Wisdom Rwanda remaining teachers/courses from workbook")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--purge-existing", action="store_true")
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 1

    entries = workbook_entries(workbook)
    db = RemoteDb(HOST, USER, PASSWORD)
    try:
        ensure_meta_columns(db)
        if args.purge_existing:
            if args.dry_run:
                print("WOULD purge prior xlsx_import course assignments/courses and remove empty temporary categories")
            else:
                db.execute_many([
                    "DELETE cr FROM course_records cr INNER JOIN courses c ON c.id = cr.course WHERE c.school_id = 27 AND (c.create_source = 'xlsx_import' OR c.code LIKE 'RB-%' OR c.code LIKE 'TV-%' OR c.code LIKE 'SP-%')",
                    "DELETE FROM courses WHERE school_id = 27 AND (create_source = 'xlsx_import' OR code LIKE 'RB-%' OR code LIKE 'TV-%' OR code LIKE 'SP-%')",
                    "DELETE cc FROM course_category cc LEFT JOIN courses c ON c.category = cc.id WHERE cc.school_id = 27 AND cc.title IN ('Secondary Subjects','TVET Subjects') AND c.id IS NULL",
                ], chunk_size=3)
        class_rows = load_classes(db)
        staff_rows = load_staff(db)
        category_cache = {row["title"]: int(row["id"]) for row in load_categories(db)}
        course_rows = load_courses(db)
        assignment_rows = load_assignments(db)

        class_map = canonical_class_map(class_rows)
        class_programs = {
            int(r["id"]): ("special" if int(r["faculty_type"] or 0) == 3 else ("tvet" if int(r["faculty_type"] or 0) == 1 else "reb"))
            for r in class_rows
        }
        password_hash = php_password_hash(DEFAULT_PASSWORD)

        staff_create = staff_update = 0
        course_create = course_update = 0
        assign_create = assign_update = assign_skip = 0
        unresolved: list[str] = []
        conflicts: list[str] = []

        staff_sql: list[str] = []
        for entry in entries:
            fname, lname = parse_person_name(entry.teacher)
            existing = None
            for row in staff_rows:
                if entry.email and normalize_email(row["email"]) == normalize_email(entry.email):
                    existing = row
                    break
                if names_match(entry.teacher, f"{row['fname']} {row['lname']}"):
                    existing = row
                    break
            phone = normalize_phone(entry.phone) or (normalize_phone(existing["phone"]) if existing else "")
            email = normalize_email(entry.email) or (normalize_email(existing["email"]) if existing else "") or f"{email_slug(fname, lname)}@wisdomschool.rw"
            if existing:
                updates = {}
                if phone and normalize_phone(existing["phone"]) != phone:
                    updates["phone"] = phone
                if email and normalize_email(existing["email"]) != email:
                    updates["email"] = email
                if not existing["post"]:
                    updates["post"] = TEACHER_POST
                if not existing["shift_id"]:
                    updates["shift_id"] = SHIFT_ID
                if updates:
                    staff_update += 1
                    if args.dry_run:
                        print(f"WOULD update staff [{existing['id']}] {fname} {lname} -> phone={phone or '-'} email={email}")
                    else:
                        staff_sql.append("UPDATE staffs SET " + ", ".join(f"{k}={sql_value(v)}" for k, v in updates.items()) + f" WHERE id={int(existing['id'])}")
            else:
                staff_create += 1
                if args.dry_run:
                    print(f"WOULD create staff {fname} {lname} phone={phone or '-'} email={email}")
                else:
                    staff_sql.append(
                        "INSERT INTO staffs (school_id,fname,lname,phone,password,status,last_login,email,post,shift_id,country,city,address,photo,lang,next_login,reset_exp,created_at,created_by,updated_at,updated_by,updateVersion) VALUES "
                        f"({SCHOOL_ID},{sql_value(fname)},{sql_value(lname)},{sql_value(phone)},{sql_value(password_hash)},2,0,{sql_value(email)},{TEACHER_POST},{SHIFT_ID},'Rwanda','Kigali','','','en',0,0,NOW(),{CREATED_BY},NOW(),{CREATED_BY},1)"
                    )

        if staff_sql and not args.dry_run:
            db.execute_many(staff_sql)
            staff_rows = load_staff(db)

        teacher_ids: dict[str, int] = {}
        staff_by_id = {int(r["id"]): r for r in staff_rows}
        for entry in entries:
            match_id = 0
            for row in staff_rows:
                if entry.email and normalize_email(row["email"]) == normalize_email(entry.email):
                    match_id = int(row["id"])
                    break
                if names_match(entry.teacher, f"{row['fname']} {row['lname']}"):
                    match_id = int(row["id"])
                    break
            if not match_id:
                raise RuntimeError(f"Could not resolve teacher after staff phase: {entry.teacher}")
            teacher_ids[entry.teacher] = match_id

        teacher_subject_loads: dict[tuple[str, str], int] = defaultdict(int)
        for entry in entries:
            title = subject_title(entry.subject)
            for _raw_class, periods in entry.classes:
                if periods > 0:
                    teacher_subject_loads[(normalize_name(entry.teacher), normalize_name(title))] += periods

        planned_candidates: list[dict[str, object]] = []
        for entry in entries:
            title = subject_title(entry.subject)
            for raw_class, periods in entry.classes:
                class_ids = classes_for_label(raw_class, entry.subject, class_map)
                if not class_ids:
                    unresolved.append(f"{entry.teacher} | {entry.subject} | {raw_class}")
                    continue
                if periods <= 0:
                    continue
                for class_id in class_ids:
                    planned_candidates.append({
                        "teacher_name": entry.teacher,
                        "title": title,
                        "credit": periods,
                        "class_id": class_id,
                        "program": class_programs.get(class_id, subject_program(entry.subject)),
                        "raw_label": raw_class,
                        "specificity": label_specificity(raw_class),
                        "subject_load": teacher_subject_loads[(normalize_name(entry.teacher), normalize_name(title))],
                    })

        if unresolved:
            for item in unresolved:
                print("UNRESOLVED", item)
            return 1

        winners: dict[tuple[str, int, str, int], dict[str, object]] = {}
        for cand in planned_candidates:
            pair_key = (
                normalize_name(str(cand["title"])),
                int(cand["credit"]),
                str(cand["program"]),
                int(cand["class_id"]),
            )
            existing_cand = winners.get(pair_key)
            if existing_cand is None:
                winners[pair_key] = cand
                continue
            chosen = choose_better_candidate(existing_cand, cand)
            if chosen is not existing_cand:
                conflicts.append(
                    f"{cand['title']} class={cand['class_id']} workbook={existing_cand['teacher_name']} ({existing_cand['raw_label']}) vs {cand['teacher_name']} ({cand['raw_label']}) -> chose {cand['teacher_name']}"
                )
                winners[pair_key] = cand
            elif existing_cand["teacher_name"] != cand["teacher_name"]:
                conflicts.append(
                    f"{cand['title']} class={cand['class_id']} workbook={existing_cand['teacher_name']} ({existing_cand['raw_label']}) vs {cand['teacher_name']} ({cand['raw_label']}) -> kept {existing_cand['teacher_name']}"
                )

        planned = [
            (
                str(cand["teacher_name"]),
                str(cand["title"]),
                int(cand["credit"]),
                int(cand["class_id"]),
                str(cand["program"]),
            )
            for cand in winners.values()
        ]

        courses_by_code, courses_by_key = build_course_maps(course_rows)
        course_sql: list[str] = []
        seen_course_keys: set[tuple[str, str, int]] = set()
        for _teacher, title, credit, _class_id, program in planned:
            key = (normalize_name(title), program, credit)
            if key in seen_course_keys:
                continue
            seen_course_keys.add(key)
            cat_id = category_cache.get(category_title(title)) or ensure_category(db, category_cache, category_title(title), args.dry_run)
            code = subject_code(title, credit, program)
            course = courses_by_code.get(code) or courses_by_key.get(key)
            if course:
                needs = course["code"] != code or normalize_name(course["title"]) != normalize_name(title) or int(float(course["marks"] or 0)) != credit * 10 or int(course["category"] or 0) != int(cat_id or 0)
                if needs:
                    course_update += 1
                    if args.dry_run:
                        print(f"WOULD update course [{course['id']}] {course['code']} -> {code} ({title}, {program}, {credit})")
                    else:
                        course_sql.append(f"UPDATE courses SET code={sql_value(code)}, title={sql_value(title)}, credit={credit}, marks={credit * 10}, program_type={sql_value(program)}, create_source={sql_value(CREATE_SOURCE)}, category={int(cat_id or 0)} WHERE id={int(course['id'])}")
            else:
                course_create += 1
                if args.dry_run:
                    print(f"WOULD create course {code} -> {title} ({program}, credit={credit})")
                else:
                    course_sql.append(f"INSERT INTO courses (school_id,title,code,category,credit,marks,program_type,create_source,created_by,updated_by) VALUES ({SCHOOL_ID},{sql_value(title)},{sql_value(code)},{int(cat_id or 0)},{credit},{credit * 10},{sql_value(program)},{sql_value(CREATE_SOURCE)},{CREATED_BY},0)")
                stub = {
                    "id": "0",
                    "code": code,
                    "title": title,
                    "credit": str(credit),
                    "marks": str(credit * 10),
                    "program_type": program,
                    "category": str(int(cat_id or 0)),
                    "create_source": CREATE_SOURCE,
                }
                courses_by_code[code] = stub
                courses_by_key[key] = stub

        if course_sql and not args.dry_run:
            db.execute_many(course_sql)
            course_rows = load_courses(db)
            courses_by_code, courses_by_key = build_course_maps(course_rows)

        assignments = {(int(r["course"]), int(r["class"])): r for r in assignment_rows}
        assign_sql: list[str] = []
        seen_assign: set[tuple[int, str, int, int]] = set()
        seen_pairs: dict[tuple[int, int], str] = {}
        for teacher_name, title, credit, class_id, program in planned:
            teacher_id = teacher_ids[teacher_name]
            code = subject_code(title, credit, program)
            course = courses_by_code.get(code) or courses_by_key.get((normalize_name(title), program, credit))
            if not course:
                raise RuntimeError(f"Missing course after course phase: {code}")
            dedupe = (teacher_id, normalize_name(title), credit, class_id)
            if dedupe in seen_assign:
                continue
            seen_assign.add(dedupe)
            pair = (int(course["id"]), class_id)
            other_teacher = seen_pairs.get(pair)
            if other_teacher and other_teacher != teacher_name:
                conflicts.append(f"{title} class={class_id} workbook={other_teacher} vs {teacher_name}")
                continue
            seen_pairs[pair] = teacher_name
            existing = assignments.get(pair)
            if existing:
                old_teacher = int(existing["lecturer"] or 0)
                same = old_teacher == teacher_id and (existing.get("term") or "") == ALL_TERMS
                if same:
                    assign_skip += 1
                    continue
                if old_teacher and old_teacher != teacher_id:
                    old = staff_by_id.get(old_teacher, {})
                    conflicts.append(f"{title} class={class_id} current={old.get('fname','')} {old.get('lname','')}".strip())
                assign_update += 1
                if args.dry_run:
                    print(f"WOULD update assignment course={course['code']} class={class_id} teacher={teacher_name}")
                else:
                    assign_sql.append(f"UPDATE course_records SET lecturer={teacher_id}, term={sql_value(ALL_TERMS)} WHERE id={int(existing['id'])}")
            else:
                assign_create += 1
                if args.dry_run:
                    print(f"WOULD assign {course['code']} -> class={class_id} teacher={teacher_name}")
                else:
                    assign_sql.append(f"INSERT INTO course_records (course,lecturer,class,year,term) VALUES ({int(course['id'])},{teacher_id},{class_id},{ACADEMIC_YEAR_ID},{sql_value(ALL_TERMS)})")

        if assign_sql and not args.dry_run:
            db.execute_many(assign_sql)

        print(f"Teachers parsed: {len(entries)}")
        print(f"Staff created={staff_create} updated={staff_update}")
        print(f"Courses created={course_create} updated={course_update}")
        print(f"Assignments created={assign_create} updated={assign_update} skipped={assign_skip}")
        if conflicts:
            print("Teacher conflicts found:")
            for item in conflicts:
                print("  " + item)
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())

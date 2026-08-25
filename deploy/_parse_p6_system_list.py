#!/usr/bin/env python3
"""Parse P6 from UPDATED LIST OF SYSTEM.xlsx (P6 sheet, or Sheet6 if that is the P6 list)."""
from __future__ import annotations

import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

SRC = Path(r"C:\methode\Desktop\methode quick docs\TREBAS\UPDATED LIST OF SYSTEM.xlsx")
OUT = Path(r"C:\xampp7\htdocs\Xander-school\deploy\_p6_system_list.json")

P6_SHEETS = {"P6"}
SHEET6_ALIASES = {"SHEET6", "SHEET 6"}
# P6 typical birth years; reject 2025/2026 placeholder dates and 2001 outliers.
DOB_MIN, DOB_MAX = 2009, 2018


def norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def excel_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        if DOB_MIN <= v.year <= DOB_MAX:
            return v.strftime("%Y-%m-%d")
        return ""
    s = str(v).strip()
    if not s or s.lower() in {"invalid date", "none", "n/a", "-"}:
        return ""
    if isinstance(v, (int, float)) and 20000 < float(v) < 60000:
        try:
            d = datetime(1899, 12, 30) + timedelta(days=int(v))
            if DOB_MIN <= d.year <= DOB_MAX:
                return d.strftime("%Y-%m-%d")
        except Exception:
            return ""
    if re.fullmatch(r"\d{4}", s):
        year = int(s)
        if DOB_MIN <= year <= DOB_MAX:
            return f"{year}-01-15"
        return ""
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if DOB_MIN <= y <= DOB_MAX and 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                datetime(y, mo, d)
            except ValueError:
                return f"{y:04d}-01-15"
            return f"{y:04d}-{mo:02d}-{d:02d}"
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            raw = s[:19] if " " in s else s
            d = datetime.strptime(raw[: len("2014-01-01") if fmt.startswith("%Y-%m-%d") and " " not in fmt else len(raw)], fmt)
            if DOB_MIN <= d.year <= DOB_MAX:
                return d.strftime("%Y-%m-%d")
            return ""
        except Exception:
            continue
    return ""


def header_map(row) -> dict[str, int]:
    mapping = {}
    for i, cell in enumerate(row):
        key = re.sub(r"[^a-z]", "", norm(cell).lower())
        if not key:
            continue
        mapping[key] = i
    return mapping


def col(row, mapping: dict[str, int], *keys) -> str:
    for key in keys:
        if key in mapping and mapping[key] < len(row):
            return norm(row[mapping[key]])
    return ""


def parse_mode(raw: str) -> str:
    s = raw.upper().replace(" ", "")
    if "BOARD" in s:
        return "0"
    if "DAY" in s:
        return "1"
    return ""


def parse_gender(raw: str) -> str:
    s = raw.strip().upper()
    if s in {"F", "FEMALE"}:
        return "F"
    if s in {"M", "MALE"}:
        return "M"
    return ""


def parse_sheet(ws, class_label: str) -> list[dict]:
    rows = ws.iter_rows(max_col=16, values_only=True)
    header = None
    mapping = {}
    students = []
    for row in rows:
        if row is None:
            continue
        joined = " ".join(norm(x) for x in row if x)
        if not joined:
            continue
        low = joined.lower().replace(" ", "")
        if header is None:
            if "fname" in low or "f.name" in low or "l.name" in low or "lname" in low:
                mapping = header_map(row)
                header = True
            continue
        fname = col(row, mapping, "fname", "firstname")
        lname = col(row, mapping, "lname", "lastname")
        if not fname and not lname:
            continue
        if fname.upper() in {"F. NAME", "NAMES", "TOTAL"}:
            continue
        excel_id = col(row, mapping, "id")
        if excel_id.endswith(".0"):
            excel_id = excel_id[:-2]
        birth_raw = row[mapping["birthyear"]] if "birthyear" in mapping and mapping["birthyear"] < len(row) else None
        students.append({
            "class_label": class_label,
            "excel_id": excel_id,
            "fname": fname,
            "lname": lname,
            "full_name": re.sub(r"\s+", " ", f"{fname} {lname}").strip(),
            "gender": parse_gender(col(row, mapping, "gender")),
            "studying_mode": parse_mode(col(row, mapping, "section", "mode")),
            "dob": excel_date(birth_raw),
            "dob_year_only": bool(re.fullmatch(r"\d{4}", str(birth_raw).strip()) if birth_raw is not None else False),
            "father": col(row, mapping, "father"),
            "ft_phone": col(row, mapping, "fathertel", "fatherphone"),
            "mother": col(row, mapping, "mother"),
            "mt_phone": col(row, mapping, "mothertel", "motherphone"),
        })
    return students


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    students = []
    skipped = []
    source_sheet = None
    names_upper = {n.strip().upper(): n for n in wb.sheetnames}
    chosen = None
    for label in P6_SHEETS:
        if label in names_upper:
            chosen = names_upper[label]
            break
    if chosen is None:
        for label, original in names_upper.items():
            if label in SHEET6_ALIASES or label.startswith("SHEET"):
                chosen = original
                break
    for name in wb.sheetnames:
        if chosen is not None and name == chosen:
            parsed = parse_sheet(wb[name], "P6")
            students.extend(parsed)
            source_sheet = name
        else:
            skipped.append(name)
    wb.close()
    payload = {
        "counts": {"P6": len(students)},
        "source_sheet": source_sheet,
        "skipped": skipped,
        "students": students,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"counts": payload["counts"], "source_sheet": source_sheet, "skipped": skipped, "total": len(students)}, indent=2))
    print("wrote", OUT)


if __name__ == "__main__":
    main()
